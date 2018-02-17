import sys
from openpyxl import Workbook
import os

dir_path = os.path.dirname(os.path.realpath(__file__))
wb = Workbook()
ws = wb.active

inputs = int(sys.argv[1])
total_entries = pow(2, inputs)

values = []
try:
    if sys.argv[2] == "-tf":
        values = ["T", "F"]
except IndexError:
    values = [0, 1] #binary by default

char_val = 65 #A
for col in range(1, inputs + 1):
    ws.cell(column=col, row=1, value=chr(char_val))
    char_val += 1

for col in range(1, inputs + 1):
    split = pow(2, inputs) / 2
    counter = 0
    val = 0
    for row in range(2, total_entries + 2):
        ws.cell(column=col, row=row, value=values[val])
        counter += 1

        if counter == split:
            val = (0 if val == 1 else 1)
            counter = 0

    inputs -= 1

wb.save(dir_path + "\\truthtable.xlsx")
